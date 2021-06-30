from PIL import Image, ImageDraw, ImageFont
import colour
from image_utils import ImageText
import piexif
import openpyxl
import os
import os.path
import pathlib


def chunk(lst, n):
	"""Yield successive n-sized chunks from lst."""
	new_list = []
	for i in range(0, len(lst), n):
		new_list.append(lst[i:i + n])
	return new_list


class Card():
	"""
    Create card to manipulate.

    size: offer list/tuple size.
    reg_font: offer regular font path.
    bold_font: offer bold font path.
    foreign_sentences: offer 10 sentences in a foreign language.
    native_sentences: offer 10 translations to the foreign langua in your native language.
    """

	def __init__(self, size, title, foreign_sentences=[], native_sentences=[], card_number=None):
		if not isinstance(size, list) and not isinstance(size, tuple):
			raise EOFError('size must be a tuple or list')
		if len(size) > 2 or len(size) < 2:
			raise EOFError('size must have 2 values xy')
		if not isinstance(foreign_sentences, list) and not isinstance(foreign_sentences, tuple):
			raise EOFError('foreign_sentences must be a tuple or list')
		if not isinstance(native_sentences, list) and not isinstance(native_sentences, tuple):
			raise EOFError('native_sentences must be a tuple or list')

		self.size = size
		self.W, self.H = size
		self.reg_font = str(pathlib.PurePath("fonts") / "JosefinSans-Regular.ttf")
		self.bold_font = str(pathlib.PurePath("fonts") / "JosefinSans-Bold.ttf")
		self.title = title
		self.card_number = card_number
		self.foreign_sentences = foreign_sentences
		self.native_sentences = native_sentences
		self.card_colors()


		# Heigh division for bboxes
		div_H1 = list(range(round(self.H / 22) * 2, self.H, round(self.H / 22)))
		if len(div_H1) > 20:
			div_H1.pop()
		div_H2 = list(range(round(self.H / 22) * 3, self.H, round(self.H / 22)))
		if len(div_H2) > 20:
			div_H2.pop()
		self.div_H1 = div_H1
		self.div_H2 = div_H2.append(self.H)
		self.div_H1_chunks = chunk(div_H1, 2)

		# Exif info.
		zeroth_ifd = {
			piexif.ImageIFD.Artist: u"Daninsky, Daniel Silva dos Santos",
			piexif.ImageIFD.Software: u"Duo Breaker script",
			piexif.ImageIFD.Copyright: u"(CC BY 4.0) Attribution 4.0 International",
			piexif.ImageIFD.XPComment: "Phrases from duolingo.com scraped with Duo Breaker script. /"
									   "@danieldaninsky, @Daninsky, @Daninsky12".encode("utf-16"),
			piexif.ImageIFD.XPAuthor: "Daninsky".encode("utf-16")
		}
		exif_dict = {"0th": zeroth_ifd}
		self.exif_bytes = piexif.dump(exif_dict)

	def card_colors(self, bg_colors=(("#545454"), ("#969696"), ("#9C9C9C")),
					font_colors=(("#FFFFFF"), ("#FFFFFF"))):
		"""
		Background colors:
		tuple of 3 colors hex:
		bg_color[0] = Background colors of the title
		bg_color[1] = Background colors of the foreign language
		bg_color[2] = Background colors of the native language

		Font colors:
		tuple of 2 colors hex:
		font_colors[0] = title font color
		font_colors[1] = sentence font color

		hex is converted here.
		"""
		self.bg_title_color = colour.hex2rgb(bg_colors[0], True)
		self.bg_first_color = colour.hex2rgb(bg_colors[1], True)
		self.bg_secon_color = colour.hex2rgb(bg_colors[2], True)

		self.font_title_color = colour.hex2rgb(font_colors[0], True)
		self.font_sente_color = colour.hex2rgb(font_colors[1], True)

	def generate_bg(self):
		"""Generate the background art of the card."""

		background = Image.new('RGB', self.size, self.bg_secon_color)
		draw = ImageDraw.Draw(background)
		title_bbox = [0, 0, self.W, round(self.H / 22) * 2]
		draw.rectangle(title_bbox, fill=self.bg_title_color)

		for chunk in self.div_H1_chunks:
			draw.rectangle(((0, chunk[0]), (self.W, chunk[1])), fill=self.bg_first_color)

		fnt = ImageFont.truetype(self.reg_font, 25)
		fnt2 = ImageFont.truetype(self.reg_font, 40)
		n_size = fnt2.getsize(str(self.card_number))[0]
		draw.text((15, 15), "@danieldaninsky", font=fnt, fill=self.font_title_color)
		if self.card_number:
			draw.text(((self.W-(30+n_size)), 30), str(self.card_number), font=fnt2, fill=self.font_title_color)


		return background

	def generate_title(self):
		"""Draw title."""

		background = self.generate_bg()
		txt = ImageText(background)
		txt.write_text_box((0, 0), self.title, box_width=self.W,
						   font_filename=self.bold_font, font_size=80,
						   color=(self.font_title_color),
						   place='center', height_compensations=(60, 0, 0))

		return background

	def generate_sentences(self):
		"""Draw sentences"""

		background = self.generate_title()
		txt = ImageText(background)
		all_sentences = []

		for sentence in zip(self.foreign_sentences, self.native_sentences):
			all_sentences.append(sentence[0])
			all_sentences.append(sentence[1])

		fnt = self.bold_font
		count = 0
		for i, sentence in enumerate(all_sentences):
			count += 1
			print(count, " - ", sentence)
			txt.write_text_box((0, self.div_H1[i]), sentence,
							   box_width=self.W,
							   font_filename=fnt,
							   font_size=42, font_min_size=34,
							   color=(self.font_sente_color),
							   place='center', height_compensations=(26, 30, 12))

			if fnt == self.bold_font:
				fnt = self.reg_font
			elif fnt == self.reg_font:
				fnt = self.bold_font

		return background

	def save(self, name):
		self.generate_sentences().save(name, exif=self.exif_bytes)


# Canvas.
size = [1080, 1920]

# Colors.
# Name convention:
# Describe the color like a class e.g. LightBlue.
# Describe the overall color or
# title color followed by the first and sec color
# separeted by underscore e.g. Green_DarkBlue_Cyan.
#(((title_fill), (foreign_fill_color), (native_fill_color)), ((title_color), (text_color)))

colors_dict = {
	"Blue":
		((("#52B1FF"), ("#156CB3"), ("#38A6FF")), (("#FFFFFF"), ("#FFFFFF"))),
	"DarkGreen":
		((("#4B8045"), ("#28801F"), ("#40CC31")), (("#FFFFFF"), ("#FFFFFF"))),
	"Pink":
		((("#FFBDE8"), ("#9F4580"), ("#FF70CD")), (("#6B2E56"), ("#FFFFFF"))),
	"DirtYellow_Purple":
		((("#AEB320"), ("#390966"), ("#6C20B3")), (("#FFFFFF"), ("#FFFFFF"))),
	"Orange":
		((("#FF7536"), ("#B33F0A"), ("#FF621B")), (("#FFFFFF"), ("#FFFFFF"))),
	"Yellow":
		((("#F2F27B"), ("#CCCC58"), ("#FFFBA0")), (("#000000"), ("#000000"))),
	"DarkBlue":
		((("#260101"), ("#033E8C"), ("#0455BF")), (("#FFFFFF"), ("#FFFFFF"))),
	"Black_DarkSlateGray_SlateGray":
		((("#000000"), ("#2F4F4F"), ("#708090")), (("#FFFFFF"), ("#FFFFFF"))),
	"Green_Blue":
		((("#448C30"), ("#3068D9"), ("#82B8D9")), (("#FFFFFF"), ("#FFFFFF"))),
	"Pink_Blue":
		((("#F25CA2"), ("#0433BF"), ("#0B9ED9")), (("#FFFFFF"), ("#FFFFFF"))),
	"Brown_Purple":
		((("#2B190E"), ("#38024D"), ("#A457BF")), (("#FFFFFF"), ("#FFFFFF"))),
	"Banana":
		((("#FFDF00"), ("#FFF249"), ("#FBFF87")), (("#402401"), ("#402401"))),
	"BananaPapayaSmoothie":
		((("#F28D77"), ("#F29A2E"), ("#F2C335")), (("#FFFFFF"), ("#FFFFFF"))),
	"DarkPurple_Orange":
		((("#150940"), ("#F29849"), ("#F2C849")), (("#FFFFFF"), ("#150940"))),
	"Orange_Lime":
		((("#F2AC29"), ("#C8F230"), ("#4FBF30")), (("#000000"), ("#000000"))),
	"Lime":
		((("#4FBF30"), ("#C8F230"), ("#4FBF30")), (("#000000"), ("#000000"))),
	"DarkBrown_DarkRed":
		((("#40342D"), ("#BF7154"), ("#D97E6A")), (("#FFFFFF"), ("#FFFFFF"))),
	"AquaGreen":
		((("#04708D"), ("#048B7E"), ("#05BF7D")), (("#FFFFFF"), ("#FFFFFF"))),
	"Gold":
		((("#BF7E04"), ("#D9A91A"), ("#F2D230")), (("#FFFFFF"), ("#592B02"))),
	"RedPassion":
		((("#8C0428"), ("#D9486E"), ("#F277A4")), (("#FFFFFF"), ("#FFFFFF"))),
	"MonoBlue":
		((("#002559"), ("#405B82"), ("#8092AC")), (("#FFFFFF"), ("#FFFFFF"))),
	"LightPink_LightPurple":
		((("#FFC4DD"), ("#F4D1FF"), ("#D1C4FF")), (("#000000"), ("#000000"))),
	"DirtGreenCamouflage":
		((("#395944"), ("#547346"), ("#85A665")), (("#FFFFFF"), ("#FFFFFF"))),
	"Brazil":
		((("#002776"), ("#009C3B"), ("#FFDF00")), (("#FFFFFF"), ("#002776"))),
	"WaterMelon":
		((("#02733E"), ("#BF0413"), ("#F21B07")), (("#FFFFFF"), ("#FFFFFF"))),
	"Forest":
		((("#733F2D"), ("#698C35"), ("#87BF34")), (("#FFFFFF"), ("#FFFFFF"))),
	"Lime2":
		((("#38A67E"), ("#6AD97B"), ("#BCF26B")), (("#FFFFFF"), ("#000000"))),
	"Dark":
		((("#05070D"), ("#101C26"), ("#283540")), (("#FFFFFF"), ("#FFFFFF"))),
	"Silver":
		((("#808080"), ("#BFBFBF"), ("#E6E6E6")), (("#000000"), ("#000000"))),
	"White":
		((("#FFFFFF"), ("#E6E6E6"), ("#FFFFFF")), (("#000000"), ("#000000"))),
	"DarkBlue_DarkCream":
		((("#203140"), ("#D9A566"), ("#F2D57E")), (("#FFFFFF"), ("#000000"))),
	"DarkCream":
		((("#BF8756"), ("#D98E32"), ("#F2D57E")), (("#000000"), ("#000000"))),
	"SpongeBob":
		((("#0457A0"), ("#AEAD0D"), ("#FFF56C")), (("#FFFFFF"), ("#000000"))),
	"MonoDarkRed":
		((("#BF7373"), ("#A63838"), ("#BF414C")), (("#FFFFFF"), ("#FFFFFF"))),
	"RedSand":
		((("#A66D6D"), ("#D99791"), ("#F2B199")), (("#000000"), ("#000000"))),
	"DarkDirt":
		((("#402B12"), ("#593F1E"), ("#8C704D")), (("#FFFFFF"), ("#FFFFFF"))),
	"Bronze":
		((("#BF978E"), ("#BF5E49"), ("#D9857E")), (("#FFFFFF"), ("#FFFFFF"))),
}


colors_list = list(colors_dict)


color = "White"

# Title
title = 'Estudos'

# Sentences
foreign = [
	"There's a group of people at the restaurant.",
	"I don't enjoy running.",
	"I like listening to guitar music.",
	"Where do you go fishing?",
	"Do you like playing soccer?",
	"Some of the girls play with dolls.",
	"Where do you go swimming?",
	"I enjoy fishing with them.",
	"A dancing group.",
	"Some of my friends go dancing every week.",
]
native = [
	"Há um grupo de pessoas no restaurante.",
	"Eu não curto correr.",
	"Eu gosto de ouvir música de violão.",
	"Onde você vai pescar?",
	"Você gosta de jogar futebol?",
	"Algumas das meninas brincam com bonecas.",
	"Onde você vai nadar?",
	"Eu curto pescar com eles.",
	"Um grupo de dança.",
	"Alguns dos meus amigos vão dançar toda semana.",
]

def render_colors_test(color_name=None, color_hex=None):
	"""
	Render colors tests, if color render one card, else, render all colors
	"""
	if color_name:
		card = Card((1080, 1920), title, foreign, native)
		card.card_colors(colors_dict[color_name][0], colors_dict[color_name][1])
		card.generate_bg()
		dir_name = pathlib.PurePath(os.path.abspath("card_colors"))
		card.save(str(dir_name
					  / '0_{}_test.jpg'.format(color_name)))
	else:
		for i, color in enumerate(colors_list, 1):
			card = Card((1080, 1920), title, foreign, native)
			card.card_colors(colors_dict[color][0], colors_dict[color][1])
			card.generate_bg()
			dir_name = pathlib.PurePath(os.path.abspath("card_colors"))
			card.save(str(dir_name
					  / '{}_{}_test.jpg'.format(i, color)))


def make_cards(file, color, title=None):
	# Setup card folder
	card_path = pathlib.Path(os.path.dirname(file)) / "cards"
	card_path.mkdir(exist_ok=True)

	card_name = os.path.basename(file)[10:-25]
	card_n = 0

	# Setup card file
	card_file = card_path / "{}_{}.jpg".format(card_name, str(card_n))

	workbook = openpyxl.open(file)
	sheetnames = workbook.sheetnames
	worksheet = workbook[sheetnames[0]]

	foreign = []
	native = []
	for row in worksheet.iter_rows(values_only=True):
		foreign.append(row[0])
		native.append(row[1])
	foreign = chunk(foreign, 10)
	native = chunk(native, 10)

	title = os.path.basename(os.path.dirname(file))

	for sentences in zip(foreign, native):
		print("=" * 50)
		card_n += 1
		card = Card((1080, 1920), title, sentences[0], sentences[1], card_number=card_n)
		card.card_colors(colors_dict[color][0], colors_dict[color][1])
		card.generate_bg()
		card.save(card_path / "{}_card_{}_{}.jpg".format(str(card_n), color, card_name))
		print("=" * 50)


if __name__ == '__main__':
	# make_cards(
	# 	r"C:\Users\daniel\EstudosLivros\English Language\Duolingo Frases\Torre 2\04 - Atividades 3\condensed_ Atividades 3.xlsx",
	# "DarkGreen", title="foda-se")
	# def test():
	# 	color_i = 0
	#
	# 	paths = sheet_utils.get_file_path(".xlsx", start_with="condensed_")
	# 	for path in paths:
	# 		print(path)
	# 		make_cards(path, colors_list[color_i])
	# 		color_i += 1

	render_colors_test(color_name=colors_list[-1])

	# for theme in basico_theme_list:
	# 	print(colors_dict[basico_theme_dict[theme]])



